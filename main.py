from bs4 import BeautifulSoup
import requests
import tkinter as tk
import openpyxl

def get_text_from_html(html_content, line_numbers):
    soup = BeautifulSoup(html_content, 'html.parser')
    lines = soup.get_text().split('\n')
    selected_lines = '\n'.join([lines[num - 1] for num in line_numbers if 0 < num <= len(lines)])
    whole_text = soup.get_text()  # Pobranie całej zawartości strony bez znaczników HTML
    
    # Znalezienie numeru linii zawierającej drugie wystąpienie słowa "Klasyki"
    klasyki_count = 0
    klasyki_line_number = 0
    for i, line in enumerate(lines):
        if "Klasyki" in line:
            klasyki_count += 1
            if klasyki_count == 2:
                klasyki_line_number = i + 1
                break
    
    return selected_lines, whole_text, klasyki_line_number

def collect_data(url):
    # Dodanie przedrostka "https://" jeśli nie jest już dodany
    if not url.startswith("http://") and not url.startswith("https://"):
        url = "https://" + url
    
    try:
        # Pobieranie zawartości strony
        response = requests.get(url, timeout=5)  # Ustawienie timeout'u na 5 sekund
        
        # Sprawdzanie, czy pobranie danych było udane
        if response.status_code == 200:
            # Pobieranie wybranych linii tekstu z zawartości strony oraz całej zawartości strony bez znaczników HTML
            selected_lines, whole_text, klasyki_line_number = get_text_from_html(response.content, [44,7,16,20,90,1])
            
            # Otwarcie lub stworzenie pliku Excel
            try:
                wb = openpyxl.load_workbook("dane.xlsx")
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                wb.active.append(["Adres URL", "Wybrane Linie", "Linijka z drugim wystąpieniem słowa 'Klasyki'", "Połączony tekst"])
            
            # Wybieranie aktywnego arkusza
            ws = wb.active
            
            # Znalezienie liczby wierszy
            max_row = ws.max_row
            
            # Dodawanie nowego wiersza z danymi
            ws.append([url, selected_lines, whole_text if klasyki_line_number != 0 else "", ""])
            
            # Dodanie formuły do komórki z połączonym tekstem
            if max_row > 1:
                ws.cell(row=max_row, column=4).value = f'=CONCATENATE(A{max_row}, C{max_row})'
            
            # Zapisywanie zmian w pliku Excel
            wb.save("dane.xlsx")
            
            status_label.config(text="Pomyślnie dodano nowe dane do pliku 'dane.xlsx'")
            print("Pomyślnie dodano nowe dane do pliku 'dane.xlsx'")
        else:
            status_label.config(text="Wystąpił błąd podczas pobierania danych ze strony")
            print("Wystąpił błąd podczas pobierania danych ze strony")
            # Otwarcie lub stworzenie pliku Excel
            try:
                wb = openpyxl.load_workbook("dane.xlsx")
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                wb.active.append(["Adres URL", "Wybrane Linie", "Linijka z drugim wystąpieniem słowa 'Klasyki'", "Połączony tekst"])
            
            # Wybieranie aktywnego arkusza
            ws = wb.active
            
            # Dodawanie nowego wiersza z danymi błędu
            ws.append([url, "Błąd", "Błąd", ""])
            
            # Dodanie formuły do komórki z połączonym tekstem
            max_row = ws.max_row
            if max_row > 1:
                ws.cell(row=max_row, column=4).value = f'=CONCATENATE(A{max_row}, C{max_row})'
            
            # Zapisywanie zmian w pliku Excel
            wb.save("dane.xlsx")
            
    except (requests.exceptions.RequestException, requests.exceptions.InvalidURL) as e:
        status_label.config(text=f"Wystąpił błąd podczas pobierania danych ze strony {url}: {e}")
        print(f"Wystąpił błąd podczas pobierania danych ze strony {url}: {e}")

def scan_from_list():
    with open("domeny.txt", 'r') as file:
        urls = file.readlines()
        urls = [url.strip() for url in urls]
        for url in urls:
            try:
                collect_data(url)
            except (requests.exceptions.RequestException, requests.exceptions.InvalidURL) as e:
                print(f"Wystąpił błąd podczas skanowania adresu {url}: {e}")
                continue

# Tworzenie głównego okna
root = tk.Tk()
root.title("Pobieranie danych")

# Ramka dla adresu URL
url_frame = tk.Frame(root)
url_frame.pack(pady=10)

# Etykieta i pole tekstowe dla adresu URL
tk.Label(url_frame, text="Wpisz Adres: ").grid(row=0, column=0)
url_entry = tk.Entry(url_frame)
url_entry.grid(row=0, column=1)

# Przycisk do zbierania danych
def collect_data_wrapper():
    url = url_entry.get()
    collect_data(url)

collect_button = tk.Button(root, text="Zbierz dane", command=collect_data_wrapper)
collect_button.pack(pady=5)

# Etykieta na status
status_label = tk.Label(root, text="")
status_label.pack(pady=5)

# Przycisk do skanowania z listy
scan_button = tk.Button(root, text="Skanuj z listy", command=scan_from_list)
scan_button.pack(pady=5)

root.mainloop()
